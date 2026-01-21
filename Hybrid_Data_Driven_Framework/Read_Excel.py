import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service


def Read_Execel():
    file = r"C:\Users\kunal\PycharmProjects\PythonProject\Automation_Testing\Data.xlsx"
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["Sheet1"]

    target_column = None
    for c in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=c).value == "ID":
            target_column = c
            break

    if target_column:
        id_value = sheet.cell(row=2, column=target_column).value
        print(f"The ID value is: {id_value}")

        # --- Selenium Logic starts here because ID was found ---
        ser_obj = Service(r"C:\Drivers\chromedriver-win64\chromedriver-win64\chromedriver.exe")
        driver = webdriver.Chrome(service=ser_obj)
        driver.get("https://the-internet.herokuapp.com/inputs")
        driver.maximize_window()
        driver.implicitly_wait(5)

        # Send the value to the input box
        driver.find_element(By.XPATH, "//input[@type='number']").send_keys(str(id_value))

        input("Press Enter to close browser...")
        driver.quit()

        return id_value  # Return the value to the main script
    else:
        print("Header 'ID' not found.")
        return None